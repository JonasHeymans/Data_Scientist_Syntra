import os
import pandas as pd
import random
from openpyxl import Workbook

# ============================================================
#   CSV BESTANDEN
# ============================================================

path = f"{os.path.abspath(os.getcwd())}/data"


# Create output folder
os.makedirs(f"{path}", exist_ok=True)

# Helper generator for dummy names
NAMES = ["Alex", "Robin", "Sam", "Jamie", "Taylor", "Casey", "Jordan", None, "Riley", "UNKNOWN"]
CITIES = ["Ghent", "Bruges", "Antwerp", "Leuven", "Hasselt", "Kortrijk", "Liège", "Namur", "Ostend", "Mechelen"]

# 0. basic
with open(f"{path}/df1.csv", "w") as f:
    f.write("name,age,score,date,value\n")
    for i in range(10):
        f.write(f"{NAMES[i]},{20+i},{50+i},{(i+1):02d}-12-2024,{100+i}\n")

# 1. delimiter='-'
with open(f"{path}/df2.csv", "w") as f:
    f.write("name;age;score\n")
    for i in range(10):
        f.write(f"{NAMES[i]};{20+i};{50+i}\n")

# 3. on_bad_lines=True (1 broken row)
with open(f"{path}/df3.csv", "w") as f:
    f.write("name,age,score\n")
    for i in range(5):
        f.write(f"{NAMES[i]},{20+i},{70+i}\n")
    f.write("THIS IS A BROKEN LINE WITHOUT COMMAS\n")
    for i in range(5,10):
        f.write(f"{NAMES[i]},{20+i},{70+i}\n")

# 4. encoding='latin1'
with open(f"{path}/df4.csv", "w", encoding="latin1") as f:
    f.write("naam,stad\n")
    special = ["José", "Anaïs", "Mårten", "Zoë", "André", "Émilie", "Renée", "Søren", "Björk", "François"]
    for i in range(10):
        f.write(f"{special[i]},{CITIES[i]}\n")

# 6. parse_dates=True + dayfirst=True
with open(f"{path}/df5.csv", "w") as f:
    f.write("date,value\n")
    for i in range(10):
        f.write(f"{(i+1):02d}/01/2025,{200+i}\n")   # dd/mm/YYYY

# 7. index_col='student'
with open(f"{path}/df6.csv", "w") as f:
    f.write("student,age,grade\n")
    for i in range(10):
        f.write(f"{NAMES[i]},{18+i},{10+i}\n")

# 8. Bonus: no header example
with open(f"{path}/df7.csv", "w") as f:
    for i in range(10):
        f.write(f"{NAMES[i]},{25+i},{80+i}\n")

print(f"CSV files generated in {path}")


# ============================================================
#   EXCEL BESTAND
#   (Alleen Excel-specifieke concepten, geen overlap met CSV)
# ============================================================


wb = Workbook()

# 0. Clean base sheet
ws1 = wb.active
ws1.title = "Sales"
ws1.append(["customer_id", "city", "amount", "date"])
for i in range(1, 11):
    ws1.append([i, f"City{i}", 100+i, f"2024-12-{i:02d}"])

# 1. Sheet with useless Excel columns (A/B/C-style)
ws2 = wb.create_sheet("UselessCols")
ws2.append(["A", "B", "C", "D", "E", "F", "G"])
for i in range(1, 11):
    ws2.append([i, i*2, i*3, "ignore", None, "REMOVE", 999+i])

# 2. Sheet with metadata at the top (skiprows concept)
ws3 = wb.create_sheet("MetadataSkip")
ws3.append(["# Report generated 2025"])
ws3.append(["# Confidential"])
ws3.append(["col1", "col2", "col3"])
for i in range(1, 11):
    ws3.append([i, i+10, i+100])


# 3. Sheet with formulas (Pandas reads stored values, not formula)
ws4 = wb.create_sheet("Formulas")
ws4.append(["A", "B", "SUM"])
for i in range(1, 11):
    ws4.append([i, i*2, f"=A{i+1}+B{i+1}"])

# Save Excel file
excel_path = f"{path}/example_excel_demo.xlsx"
wb.save(excel_path)

print(f"Excel demo file generated in {excel_path}")
